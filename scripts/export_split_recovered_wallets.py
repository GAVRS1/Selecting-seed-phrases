#!/usr/bin/env python3
"""Export rows from split recovered_wallets_* tables to XLSX and then delete them.

Behavior:
1. Read all rows from BTC/EVM/SOL tables in batches (ordered by id inside each table).
2. Build an Excel workbook with sheet "all" + grouped sheets (btc/evm/sol/unknown).
3. If (and only if) export succeeds, delete only exported rows from corresponding source tables.

If export fails, no deletions are performed.
"""

from __future__ import annotations

import argparse
import os
import re
import subprocess
import sys
import tempfile
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from typing import Iterable

PSQL_BIN = os.environ.get("PSQL_BIN", "psql")


@dataclass(frozen=True)
class SplitRow:
    source_table: str
    row_id: int
    created_at: str
    blockchain: str
    address: str
    mnemonic: str


def parse_env_file(path: str) -> dict[str, str]:
    env: dict[str, str] = {}
    if not os.path.isfile(path):
        return env

    with open(path, "r", encoding="utf-8") as handle:
        for raw_line in handle:
            line = raw_line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            env[key.strip()] = value.strip().strip("\"'")
    return env


def resolve_connection(cli_conn: str | None, env_file: str) -> str:
    env_values = parse_env_file(env_file)
    conn = cli_conn or env_values.get("RECOVERY_POSTGRES_CONN") or os.environ.get("RECOVERY_POSTGRES_CONN")
    if not conn:
        raise ValueError(
            "PostgreSQL connection string is missing. Use --postgres-conn or set RECOVERY_POSTGRES_CONN in .env/environment."
        )
    return conn


def validate_table_name(name: str) -> str:
    if not re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]*", name):
        raise ValueError(f"Invalid PostgreSQL table name: {name}")
    return name


def run_psql(conn: str, sql: str) -> str:
    """Run SQL via a temp file to avoid Windows command-length limits."""
    with tempfile.NamedTemporaryFile("w", suffix=".sql", delete=False, encoding="utf-8") as handle:
        handle.write(sql)
        sql_path = handle.name

    cmd = [PSQL_BIN, conn, "-v", "ON_ERROR_STOP=1", "-At", "-F", "\t", "-f", sql_path]
    result = subprocess.run(cmd, capture_output=True, text=True, check=False)
    try:
        os.unlink(sql_path)
    except OSError:
        pass

    if result.returncode != 0:
        stderr = result.stderr or ""
        stdout = result.stdout or ""
        raise RuntimeError(f"psql failed: {stderr.strip() or stdout.strip()}")
    return result.stdout or ""


def fetch_rows(conn: str, source_table: str, batch_size: int, after_id: int = 0) -> list[SplitRow]:
    where_clause = f"WHERE id > {after_id} " if after_id > 0 else ""
    sql = (
        f"SELECT id, created_at, blockchain, address, mnemonic FROM {source_table} "
        f"{where_clause}ORDER BY id LIMIT {batch_size};"
    )
    output = run_psql(conn, sql)

    rows: list[SplitRow] = []
    for line in output.splitlines():
        if not line.strip():
            continue
        parts = line.split("\t")
        if len(parts) != 5:
            raise RuntimeError(f"Unexpected row format from psql: {line!r}")
        rows.append(
            SplitRow(
                source_table=source_table,
                row_id=int(parts[0]),
                created_at=parts[1].strip(),
                blockchain=parts[2].strip(),
                address=parts[3].strip(),
                mnemonic=parts[4].strip(),
            )
        )
    return rows


def classify_chain(blockchain: str) -> str | None:
    chain = blockchain.strip().lower()
    if chain in {"btc", "bitcoin"}:
        return "btc"
    if chain in {"eth", "ethereum", "evm", "bsc", "polygon", "arbitrum", "optimism", "avalanche", "base"}:
        return "evm"
    if chain in {"sol", "solana"}:
        return "sol"
    return None


def export_xlsx(rows: list[SplitRow], output_path: str) -> None:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required for Excel export. Install it with: python3 -m pip install openpyxl"
        ) from exc

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    groups: dict[str, list[SplitRow]] = defaultdict(list)
    for row in rows:
        mapped = classify_chain(row.blockchain)
        sheet_key = mapped if mapped is not None else "unknown"
        groups[sheet_key].append(row)

    all_sheet = wb.create_sheet("all")
    header = ["source_table", "id", "created_at", "source_blockchain", "group_chain", "address", "mnemonic"]
    all_sheet.append(header)
    for row in rows:
        all_sheet.append([
            row.source_table,
            row.row_id,
            row.created_at,
            row.blockchain,
            classify_chain(row.blockchain) or "unknown",
            row.address,
            row.mnemonic,
        ])

    for sheet_name in ("btc", "evm", "sol", "unknown"):
        sheet = wb.create_sheet(sheet_name)
        sheet.append(header)
        for row in groups.get(sheet_name, []):
            sheet.append([
                row.source_table,
                row.row_id,
                row.created_at,
                row.blockchain,
                classify_chain(row.blockchain) or "unknown",
                row.address,
                row.mnemonic,
            ])

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)


def chunked_ids(rows: list[SplitRow], chunk_size: int = 5000) -> Iterable[list[int]]:
    ids = [row.row_id for row in rows]
    for start in range(0, len(ids), chunk_size):
        yield ids[start : start + chunk_size]


def delete_exported_rows(conn: str, source_table: str, rows: list[SplitRow]) -> int:
    deleted_total = 0
    for id_chunk in chunked_ids(rows):
        sql = f"""
WITH deleted AS (
    DELETE FROM {source_table}
    WHERE id IN ({",".join(str(row_id) for row_id in id_chunk)})
    RETURNING id
)
SELECT COUNT(*) FROM deleted;
"""
        output = run_psql(conn, sql).strip()
        if not output or not output.isdigit():
            raise RuntimeError(f"Unable to parse delete count for chunk in {source_table}: {output!r}")
        deleted_total += int(output)
    return deleted_total


def parse_args(argv: Iterable[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export rows from recovered_wallets_btc/evm/sol to Excel and delete exported rows from source tables."
    )
    parser.add_argument("--env-file", default=".env", help="Path to .env file (default: .env)")
    parser.add_argument("--postgres-conn", default=None, help="PostgreSQL connection string")
    parser.add_argument("--table-btc", default="recovered_wallets_btc", help="BTC source table")
    parser.add_argument("--table-evm", default="recovered_wallets_evm", help="EVM source table")
    parser.add_argument("--table-sol", default="recovered_wallets_sol", help="SOL source table")
    parser.add_argument("--batch-size", type=int, default=1000, help="Rows to fetch per DB batch (per table)")
    parser.add_argument(
        "--excel-output",
        default=f"split_wallets_export_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx",
        help="Path to output XLSX file",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Only export XLSX and print stats; do not delete rows from DB.",
    )
    return parser.parse_args(argv)


def main(argv: Iterable[str] | None = None) -> int:
    args = parse_args(argv)
    if args.batch_size <= 0:
        print("--batch-size must be > 0", file=sys.stderr)
        return 2

    try:
        conn = resolve_connection(args.postgres_conn, args.env_file)
        source_tables = {
            "btc": validate_table_name(args.table_btc),
            "evm": validate_table_name(args.table_evm),
            "sol": validate_table_name(args.table_sol),
        }

        all_rows: list[SplitRow] = []
        stats: dict[str, int] = {"btc": 0, "evm": 0, "sol": 0, "unknown": 0}

        for chain_name, source_table in source_tables.items():
            first_batch = fetch_rows(conn, source_table, args.batch_size)
            if not first_batch:
                print(f"Table {source_table}: no rows found.")
                continue

            table_rows: list[SplitRow] = []
            batch_index = 0
            next_after_id = 0

            while True:
                rows = first_batch if batch_index == 0 else fetch_rows(conn, source_table, args.batch_size, next_after_id)
                if not rows:
                    break

                batch_index += 1
                table_rows.extend(rows)
                all_rows.extend(rows)
                for row in rows:
                    stats[classify_chain(row.blockchain) or "unknown"] += 1

                print(
                    f"Table {source_table}, batch {batch_index}: fetched {len(rows)} rows "
                    f"(ids {rows[0].row_id}..{rows[-1].row_id})"
                )
                next_after_id = rows[-1].row_id

            print(f"Table {source_table}: total fetched {len(table_rows)} rows.")

        if not all_rows:
            print("No rows found in split wallet tables; nothing to export.")
            return 0

        export_xlsx(all_rows, args.excel_output)
        print(f"Exported {len(all_rows)} rows to Excel: {args.excel_output}")
        print(f"Split stats => btc={stats['btc']} evm={stats['evm']} sol={stats['sol']} unknown={stats['unknown']}")

        if args.dry_run:
            print("Dry-run enabled: deletion skipped.")
            return 0

        deleted_by_table: dict[str, int] = {}
        for source_table in source_tables.values():
            rows_for_table = [row for row in all_rows if row.source_table == source_table]
            if not rows_for_table:
                deleted_by_table[source_table] = 0
                continue
            deleted_by_table[source_table] = delete_exported_rows(conn, source_table, rows_for_table)

        deleted_total = sum(deleted_by_table.values())
        print(
            "Deleted rows after successful export: "
            + ", ".join(f"{table}={count}" for table, count in deleted_by_table.items())
            + f" (total={deleted_total})"
        )
        return 0
    except Exception as exc:  # noqa: BLE001
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
