#!/usr/bin/env python3
"""Migrate rows from legacy recovered_wallets table to split tables and export XLSX.

Behavior:
1. Read rows from legacy table in batches (ordered by id).
2. Build an Excel workbook with one sheet per blockchain group.
3. Insert mnemonics into seed_phrases_{btc|evm|sol}.
4. Insert wallet rows into recovered_wallets_{btc|evm|sol}.
5. Delete only processed legacy rows in the same SQL transaction.

If DB write transaction fails, no deletions are performed.
"""

from __future__ import annotations

import argparse
import csv
import os
import re
import subprocess
import sys
import tempfile
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from typing import Iterable

PSQL_BIN = os.environ.get("PSQL_BIN", "psql")


@dataclass(frozen=True)
class LegacyRow:
    row_id: int
    created_at: str
    blockchain: str
    address: str
    mnemonic: str


def parse_env_file(path: str) -> dict[str, str]:
    env: dict[str, str] = {}
    if not os.path.isfile(path):
        return env

    with open(path, "r", encoding="utf-8") as handle:
        for raw_line in handle:
            line = raw_line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            env[key.strip()] = value.strip().strip("\"'")
    return env


def resolve_connection(cli_conn: str | None, env_file: str) -> str:
    env_values = parse_env_file(env_file)
    conn = cli_conn or env_values.get("RECOVERY_POSTGRES_CONN") or os.environ.get("RECOVERY_POSTGRES_CONN")
    if not conn:
        raise ValueError(
            "PostgreSQL connection string is missing. Use --postgres-conn or set RECOVERY_POSTGRES_CONN in .env/environment."
        )
    return conn


def validate_table_name(name: str) -> str:
    if not re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]*", name):
        raise ValueError(f"Invalid PostgreSQL table name: {name}")
    return name


def sql_literal(value: str) -> str:
    return "'" + value.replace("'", "''") + "'"


def run_psql(conn: str, sql: str) -> str:
    cmd = [PSQL_BIN, conn, "-v", "ON_ERROR_STOP=1", "-At", "-F", "\t", "-c", sql]
    result = subprocess.run(cmd, capture_output=True, text=True, check=False)
    if result.returncode != 0:
        raise RuntimeError(f"psql failed: {result.stderr.strip() or result.stdout.strip()}")
    return result.stdout


def run_psql_file(conn: str, sql_file: str) -> str:
    cmd = [PSQL_BIN, conn, "-v", "ON_ERROR_STOP=1", "-At", "-F", "\t", "-f", sql_file]
    result = subprocess.run(cmd, capture_output=True, text=True, check=False)
    if result.returncode != 0:
        raise RuntimeError(f"psql failed: {result.stderr.strip() or result.stdout.strip()}")
    return result.stdout


def fetch_legacy_rows(conn: str, legacy_table: str, batch_size: int) -> list[LegacyRow]:
    sql = (
        f"SELECT id, created_at, blockchain, address, mnemonic FROM {legacy_table} "
        f"ORDER BY id LIMIT {batch_size};"
    )
    output = run_psql(conn, sql)

    rows: list[LegacyRow] = []
    for line in output.splitlines():
        if not line.strip():
            continue
        parts = line.split("\t")
        if len(parts) != 5:
            raise RuntimeError(f"Unexpected row format from psql: {line!r}")
        rows.append(
            LegacyRow(
                row_id=int(parts[0]),
                created_at=parts[1].strip(),
                blockchain=parts[2].strip(),
                address=parts[3].strip(),
                mnemonic=parts[4].strip(),
            )
        )
    return rows


def classify_chain(blockchain: str) -> str | None:
    chain = blockchain.strip().lower()
    if chain in {"btc", "bitcoin"}:
        return "btc"
    if chain in {"eth", "ethereum", "evm", "bsc", "polygon", "arbitrum", "optimism", "avalanche", "base"}:
        return "evm"
    if chain in {"sol", "solana"}:
        return "sol"
    return None


def write_batch_tsv(rows: list[LegacyRow], output_path: str) -> None:
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    with open(output_path, "w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(
            handle,
            delimiter="\t",
            quotechar='"',
            quoting=csv.QUOTE_MINIMAL,
            lineterminator="\n",
        )
        for row in rows:
            writer.writerow([row.row_id, row.blockchain, row.address, row.mnemonic])


def export_xlsx(rows: list[LegacyRow], output_path: str) -> None:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required for Excel export. Install it with: python3 -m pip install openpyxl"
        ) from exc

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    groups: dict[str, list[LegacyRow]] = defaultdict(list)
    for row in rows:
        mapped = classify_chain(row.blockchain)
        sheet_key = mapped if mapped is not None else "unknown"
        groups[sheet_key].append(row)

    all_sheet = wb.create_sheet("all")
    header = ["id", "created_at", "source_blockchain", "group_chain", "address", "mnemonic"]
    all_sheet.append(header)
    for row in rows:
        all_sheet.append([
            row.row_id,
            row.created_at,
            row.blockchain,
            classify_chain(row.blockchain) or "unknown",
            row.address,
            row.mnemonic,
        ])

    for sheet_name in ("btc", "evm", "sol", "unknown"):
        sheet = wb.create_sheet(sheet_name)
        sheet.append(header)
        for row in groups.get(sheet_name, []):
            sheet.append([
                row.row_id,
                row.created_at,
                row.blockchain,
                classify_chain(row.blockchain) or "unknown",
                row.address,
                row.mnemonic,
            ])

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)


def to_psql_path(path: str) -> str:
    return path.replace("\\", "/")


def build_transfer_sql(legacy_table: str, batch_file_path: str) -> str:
    batch_file_literal = sql_literal(to_psql_path(batch_file_path))
    return f"""
BEGIN;
CREATE TEMP TABLE tmp_batch (
    id BIGINT NOT NULL,
    blockchain TEXT NOT NULL,
    address TEXT NOT NULL,
    mnemonic TEXT NOT NULL
) ON COMMIT DROP;

\\copy tmp_batch (id, blockchain, address, mnemonic) FROM {batch_file_literal} WITH (FORMAT csv, DELIMITER E'\\t', QUOTE '\"', ESCAPE '\"');

WITH classified AS (
    SELECT
        id,
        blockchain,
        address,
        mnemonic,
        CASE
            WHEN lower(blockchain) IN ('btc', 'bitcoin') THEN 'btc'
            WHEN lower(blockchain) IN ('eth', 'ethereum', 'evm', 'bsc', 'polygon', 'arbitrum', 'optimism', 'avalanche', 'base') THEN 'evm'
            WHEN lower(blockchain) IN ('sol', 'solana') THEN 'sol'
            ELSE NULL
        END AS target_chain
    FROM tmp_batch
)
INSERT INTO seed_phrases_btc (mnemonic)
SELECT DISTINCT mnemonic FROM classified WHERE target_chain = 'btc'
ON CONFLICT (mnemonic) DO NOTHING;

INSERT INTO seed_phrases_evm (mnemonic)
SELECT DISTINCT mnemonic FROM classified WHERE target_chain = 'evm'
ON CONFLICT (mnemonic) DO NOTHING;

INSERT INTO seed_phrases_sol (mnemonic)
SELECT DISTINCT mnemonic FROM classified WHERE target_chain = 'sol'
ON CONFLICT (mnemonic) DO NOTHING;

INSERT INTO recovered_wallets_btc (blockchain, address, mnemonic)
SELECT blockchain, address, mnemonic FROM classified WHERE target_chain = 'btc'
ON CONFLICT (blockchain, address, mnemonic) DO NOTHING;

INSERT INTO recovered_wallets_evm (blockchain, address, mnemonic)
SELECT blockchain, address, mnemonic FROM classified WHERE target_chain = 'evm'
ON CONFLICT (blockchain, address, mnemonic) DO NOTHING;

INSERT INTO recovered_wallets_sol (blockchain, address, mnemonic)
SELECT blockchain, address, mnemonic FROM classified WHERE target_chain = 'sol'
ON CONFLICT (blockchain, address, mnemonic) DO NOTHING;

DELETE FROM {legacy_table}
WHERE id IN (SELECT id FROM tmp_batch);
COMMIT;
"""


def parse_args(argv: Iterable[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Migrate legacy recovered_wallets rows to split tables with Excel export and delete processed rows."
    )
    parser.add_argument("--env-file", default=".env", help="Path to .env file (default: .env)")
    parser.add_argument("--postgres-conn", default=None, help="PostgreSQL connection string")
    parser.add_argument("--legacy-table", default="recovered_wallets", help="Legacy source table")
    parser.add_argument("--batch-size", type=int, default=1000, help="Rows to process in one run")
    parser.add_argument(
        "--excel-output",
        default=f"legacy_wallets_export_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx",
        help="Path to output XLSX file",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Only export XLSX and print stats; do not modify DB and do not delete rows.",
    )
    return parser.parse_args(argv)


def main(argv: Iterable[str] | None = None) -> int:
    args = parse_args(argv)
    if args.batch_size <= 0:
        print("--batch-size must be > 0", file=sys.stderr)
        return 2

    try:
        conn = resolve_connection(args.postgres_conn, args.env_file)
        legacy_table = validate_table_name(args.legacy_table)
        rows = fetch_legacy_rows(conn, legacy_table, args.batch_size)

        if not rows:
            print("No rows found in legacy table; nothing to migrate.")
            return 0

        export_xlsx(rows, args.excel_output)

        stats: dict[str, int] = {"btc": 0, "evm": 0, "sol": 0, "unknown": 0}
        for row in rows:
            stats[classify_chain(row.blockchain) or "unknown"] += 1

        print(f"Exported {len(rows)} rows to Excel: {args.excel_output}")
        print(f"Split stats => btc={stats['btc']} evm={stats['evm']} sol={stats['sol']} unknown={stats['unknown']}")

        if args.dry_run:
            print("Dry-run enabled: DB transfer and deletion skipped.")
            return 0

        with tempfile.TemporaryDirectory(prefix="legacy_wallet_migration_") as temp_dir:
            batch_file_path = os.path.join(temp_dir, "batch.tsv")
            transfer_sql_path = os.path.join(temp_dir, "transfer.sql")

            write_batch_tsv(rows, batch_file_path)
            transfer_sql = build_transfer_sql(legacy_table, batch_file_path)
            with open(transfer_sql_path, "w", encoding="utf-8", newline="\n") as handle:
                handle.write(transfer_sql)

            run_psql_file(conn, transfer_sql_path)

        print(
            "Transferred and deleted rows from legacy table: "
            f"{len(rows)} (ids {rows[0].row_id}..{rows[-1].row_id})"
        )
        return 0
    except Exception as exc:  # noqa: BLE001
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
